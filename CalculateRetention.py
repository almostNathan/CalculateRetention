from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from tkinter import filedialog
import os


#get the workbook filepath from the user, dialog fileselect window
workbookFilePath = filedialog.askopenfilename(initialdir="C:/", title="Select the Excel File", filetypes=(("Excel Files", "*.xlsx"),("All Files", "*.*")))
#open the workbook
workbook = load_workbook(workbookFilePath)
#get the active sheet
sheet = workbook.active
#get the maximum rows
maxRows = sheet.max_row

writtenOffA = sheet[f"A{maxRows+1}"]
writtenOffB = sheet[f"B{maxRows+1}"]
writtenOffC = sheet[f"C{maxRows+1}"]

collectedA = sheet[f"A{maxRows+2}"]
collectedB = sheet[f"B{maxRows+2}"]
collectedC = sheet[f"C{maxRows+2}"]

inCollectionsA = sheet[f"A{maxRows+3}"]
inCollectionsB = sheet[f"B{maxRows+3}"]
inCollectionsC = sheet[f"C{maxRows+3}"]

totalA = sheet[f"A{maxRows+4}"]
totalB = sheet[f"B{maxRows+4}"]
totalC = sheet[f"C{maxRows+4}"]

writtenOffA.border = Border(top=Side(style="thin"),left=Side(style="thin"))
writtenOffB.border = Border(top=Side(style="thin"))
writtenOffC.border = Border(top=Side(style="thin"),right=Side(style="thin"))

collectedA.border = Border(left=Side(style="thin"))
collectedC.border = Border(right=Side(style="thin"))

inCollectionsA.border = Border(left=Side(style="thin"))
inCollectionsC.border = Border(right=Side(style="thin"))

totalA.border = Border(bottom=Side(style="thin"),left=Side(style="thin"))
totalB.border = Border(bottom=Side(style="thin"))
totalC.border = Border(bottom=Side(style="thin"),right=Side(style="thin"))

#written Off
writtenOffA.value = "Written Off"
writtenOffB.value = f"=SUMIFS(I2:I{maxRows},F2:F{maxRows},FALSE)"
#percent written off
writtenOffC.value = f"=B{maxRows+1}/B{maxRows+4}"
writtenOffC.number_format = "0.00%"

#Retained
collectedA.value = "Collected"
collectedB.value = f"=SUMIFS(I2:I{maxRows},F2:F{maxRows},TRUE,K2:K{maxRows},\"=0\")"
#percent Retained(collected)
collectedC.value = f"=B{maxRows+2}/B{maxRows+4}"
collectedC.number_format = "0.00%"

#AB Due (in Collections)
inCollectionsA.value = "In Collections"
inCollectionsB.value = f"=SUMIFS(I2:I{maxRows},F2:F{maxRows},TRUE,K2:K{maxRows},\"<>0\")"
#percent AB Due
inCollectionsC.value = f"=B{maxRows+3}/B{maxRows+4}"
inCollectionsC.number_format = "0.00%"

#BOLD the inCollections Row
inCollectionsA.font = Font(bold = True)
inCollectionsB.font = Font(bold = True)
inCollectionsC.font = Font(bold = True)

#total Billed
totalA.value = "Total Billed"
totalB.value = f"=SUM(I2:I{maxRows})"


workbook.save(workbookFilePath)
workbook.close

os.startfile(workbookFilePath)
